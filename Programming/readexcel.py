import openpyxl

count = 0
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("hello_world.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(3, dataframe1.max_column):
        print(col[row].value, end=" ")
        if int(col[row].value) > 80:
            count = count +1

print(count)

